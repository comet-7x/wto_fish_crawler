"""Build the SITE deliverable: every page + file reachable from the fish_e portal.

Reads a Tier-1 crawl (manifest.jsonl + markdown/) and produces:
  * 渔业补贴站点库/  organised  类型 -> 类别 :
        01_网页_HTML转Markdown/<中文类别>/<名称>.md
        02_文件_PDF/<中文类别>/<名称>.pdf
        03_音视频_受限仅链接/_说明.txt
  * 渔业补贴站点库/_站点文件索引.{xlsx,csv}  — one row per page/file with a
    clickable original URL, a real title (anchor text from the linking page,
    not the generic <title>), the parent page, category, local path, status.

Run:
    python tools/build_site_library.py --out ./site_out
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import re
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CATEGORY_ZH = {
    "overview": "概览", "introduction": "导论", "legal_text": "法律文本",
    "ratification": "接受与批准", "implementation": "履约", "fish_fund": "渔业基金",
    "publication": "出版物", "news": "新闻", "ministerial": "部长会简报",
    "case_story": "案例故事", "international_instrument": "国际文书",
    "negotiation_submission": "谈判", "multimedia": "音视频", "uncategorized": "未分类",
}
# Curated Chinese names (reused from the checklist) for the well-known files.
NAME_ZH = {
    "fish_e.htm": "渔业补贴专题门户·协定概览",
    "agreement_fisheries_subsidies_e.htm": "如何接受议定书（成员程序指引）",
    "fish_acceptances_e.htm": "成员接受·批准清单",
    "implementfishagreement22_e.htm": "发展中·最不发达国家履约报告",
    "fish_fund_e.htm": "WTO 渔业基金（技援与能力建设）",
    "fishagree_e.htm": "协定宣传册（目标与预期收益）",
    "fish_factsheet_e.pdf": "渔业补贴情况说明书（Factsheet）",
    "fish_15sep25_e.htm": "协定生效新闻（2025-09-15）",
    "fish_10oct22_e.htm": "渔业补贴新闻（2022-10-10）",
    "fish_08nov22_e.htm": "渔业补贴新闻（2022-11-08）",
    "fish_14jun22_e.htm": "渔业补贴新闻（2022-06-14）",
    "fisheries_subsidies_e.htm": "MC13 部长会渔业补贴简报",
    "bffish_e.htm": "MC12 部长会渔业补贴简报",
    "stories_tanzania_e.htm": "案例·坦桑尼亚",
    "stories-fijipacific_e.htm": "案例·斐济/太平洋",
    "stories_from_the_ocean_e.htm": "案例·海洋故事(总览)",
    "stories-trinidadtobago_e.htm": "案例·特立尼达和多巴哥",
    "stories_oman_e.htm": "案例·阿曼", "stories-malaysia_e.htm": "案例·马来西亚",
    "1969_vclt.pdf": "1969 维也纳条约法公约（VCLT）",
    "1982_unclos.pdf": "1982 联合国海洋法公约（UNCLOS）",
    "1995_fao_ccrf.pdf": "1995 FAO 负责任渔业行为守则（CCRF）",
    "1995_unfsa.pdf": "1995 联合国鱼类种群协定（UNFSA）",
    "2001_ipoa_iuu.pdf": "2001 FAO IUU 行动计划（IPOA-IUU）",
    "2009_psma.pdf": "2009 港口国措施协定（PSMA）",
    "2014_vg_fsp.pdf": "2014 FAO 船旗国表现自愿准则（VG-FSP）",
    "2015_fao_ssf.pdf": "2015 FAO 小规模渔业自愿准则（SSF）",
    "2017_vg_cds.pdf": "2017 FAO 渔获文件方案自愿准则（VG-CDS）",
    "ngr_presentation_on_rfmo_as.pdf": "规则谈判组 RFMO/As 演示材料",
    "implementfishagreement22_e.pdf": "实施《渔业补贴协定》出版物（完整版）",
    "fishagree_e.pdf": "协定宣传册（PDF 版）",
    "2024_07_05_fish_chair_update.mp4": "主席渔业谈判进展视频（2024-07-05）",
    "fish_20240222.mp4": "MC13 渔业补贴视频（2024-02-22）",
    "fish_28022024.mp4": "MC13 渔业补贴视频（2024-02-28）",
}
NAME_ZH_URL = {"/docs_e/legal_e/fish_e.htm": "渔业补贴协定法律文本（正式条文）"}
ROOT = Path("渔业补贴站点库")
_TRIVIAL = {"here", "more", "pdf", "english", "back to top"}


def _basename(url: str) -> str:
    return url.rstrip("/").split("?", 1)[0].split("/")[-1]


def _safe(name: str) -> str:
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip()[:120]


def anchor_map(out: Path) -> dict[str, str]:
    """url -> first descriptive anchor text seen in any crawled markdown."""
    amap: dict[str, str] = {}
    for f in glob.glob(str(out / "markdown" / "*.md")):
        for m in re.finditer(r"\[([^\]]+)\]\((https?://[^)]+)\)",
                              Path(f).read_text(encoding="utf-8")):
            t, u = m.group(1).strip(), m.group(2).split("#")[0]
            if t and not t.startswith("http"):
                amap.setdefault(u, t)
    return amap


def title_en(url: str, amap: dict, page_title: str | None) -> str:
    a = amap.get(url) or amap.get(url.replace("https://www.wto.org", ""))
    if a and len(a) > 3 and a.lower() not in _TRIVIAL and not re.fullmatch(r"[\d\s.–-]+", a):
        return a
    if page_title and page_title.strip().upper() != "WORLD TRADE ORGANIZATION":
        return page_title.strip()
    return ""


def name_zh(url: str, base: str) -> str:
    for sub, n in NAME_ZH_URL.items():
        if sub in url:
            return n
    return NAME_ZH.get(base, "")


def main() -> int:
    ap = argparse.ArgumentParser(description="Build the fish_e site deliverable")
    ap.add_argument("--out", required=True, help="Tier-1 crawl dir (manifest.jsonl, markdown/)")
    args = ap.parse_args()
    out = Path(args.out)
    rows = [json.loads(l) for l in (out / "manifest.jsonl").read_text(encoding="utf-8").splitlines()
            if l.strip()]
    rows = [r for r in rows if not r.get("duplicate_of")]
    amap = anchor_map(out)

    if ROOT.exists():
        shutil.rmtree(ROOT, ignore_errors=True)
    items = []
    for r in rows:
        url = r["url"]
        base = _basename(url)
        blocked = bool(r.get("error") and "access-gated" in (r.get("error") or ""))
        if r.get("error") and not blocked:          # soft-404 / failures: skip
            continue
        cat = r.get("category", "uncategorized")
        zh_cat = CATEGORY_ZH.get(cat, cat)
        nm_zh = name_zh(url, base)
        # For pages prefer the real H1 from the Markdown over the generic <title>.
        page_title = r.get("title")
        if r.get("out_md_path"):
            md = (out / r["out_md_path"]).read_text(encoding="utf-8")
            h1 = re.search(r"^#\s+(.+)$", md.split("---", 2)[-1], re.M)
            if h1:
                page_title = h1.group(1).strip()
        t_en = title_en(url, amap, page_title)
        disp = nm_zh or t_en or base.rsplit(".", 1)[0]
        if r.get("out_md_path"):
            kind, top, src, ext = "网页", "01_网页_HTML转Markdown", out / r["out_md_path"], "md"
            status = "已转 Markdown"
        elif r.get("content_type") == "pdf":
            kind, top, src, ext = "PDF", "02_文件_PDF", out / f"raw/pdf/{r['raw_sha256']}.pdf", "pdf"
            status = "已下载·待解析"
        elif blocked:
            kind, top, src, ext = "音视频", "03_音视频_受限仅链接", None, base.rsplit(".", 1)[-1]
            status = "受限·需浏览器获取"
        else:
            continue
        local = ""
        if src and src.exists():
            folder = ROOT / top / zh_cat
            folder.mkdir(parents=True, exist_ok=True)
            dest = folder / f"{_safe(disp)}.{ext}"
            dest.write_bytes(src.read_bytes())
            local = str(dest).replace("\\", "/")
        items.append({
            "类型": kind, "类别": zh_cat, "名称(中文)": nm_zh, "英文标题": t_en,
            "父页面": r.get("source_url") or "(种子/门户)", "原始URL": url,
            "本地路径": local, "状态": status,
        })

    # media note
    media = [it for it in items if it["类型"] == "音视频"]
    if media:
        (ROOT / "03_音视频_受限仅链接").mkdir(parents=True, exist_ok=True)
        (ROOT / "03_音视频_受限仅链接" / "_说明.txt").write_text(
            "以下视频受访问限制，无法直接下载，仅留官方链接：\n\n"
            + "\n".join(f"- {it['英文标题'] or it['原始URL']}\n  {it['原始URL']}" for it in media)
            + "\n", encoding="utf-8")

    order = ["网页", "PDF", "音视频"]
    items.sort(key=lambda it: (order.index(it["类型"]), it["类别"], it["名称(中文)"] or it["英文标题"]))

    # CSV
    cols = ["序号", "类型", "类别", "名称(中文)", "英文标题", "父页面", "原始URL", "本地路径", "状态"]
    with (ROOT / "_站点文件索引.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f); w.writerow(cols)
        for i, it in enumerate(items, 1):
            w.writerow([i] + [it[c] for c in cols[1:]])
    # XLSX
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "站点文件索引"
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="4472C4")
    for i, it in enumerate(items, 1):
        ws.append([i, it["类型"], it["类别"], it["名称(中文)"], it["英文标题"],
                   it["父页面"], it["原始URL"], it["本地路径"], it["状态"]])
        lk = ws.cell(ws.max_row, 7)
        lk.hyperlink = it["原始URL"]; lk.font = Font(color="0563C1", underline="single")
        ws.cell(ws.max_row, 5).alignment = Alignment(wrap_text=False)
        if it["类型"] == "音视频":
            for c in ws[ws.max_row]:
                c.fill = PatternFill("solid", fgColor="FCE4E4")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:I{ws.max_row}"
    for col, w_ in zip("ABCDEFGHI", [5, 8, 12, 34, 60, 46, 60, 50, 16]):
        ws.column_dimensions[col].width = w_
    wb.save(ROOT / "_站点文件索引.xlsx")

    n_html = sum(1 for it in items if it["类型"] == "网页")
    n_pdf = sum(1 for it in items if it["类型"] == "PDF")
    print(f"站点库: {len(items)} 条 (网页 {n_html} / PDF {n_pdf} / 音视频 {len(media)})")
    print(f"  {ROOT}/  + _站点文件索引.xlsx/.csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
