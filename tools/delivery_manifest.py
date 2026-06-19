"""Per-file delivery manifest: every in-scope file with a provable URL + reason.

Built to satisfy two delivery requirements:
  1. every downloaded / to-be-downloaded file carries an ORIGINAL URL that proves
     the document actually exists (openable on wto.org / docs.wto.org);
  2. every file we decided to keep carries a concrete REASON (for fisheries
     documents, the exact title keyword that matched — not a vague "relevant").

Outputs (under docs_manifest/):
  * delivery_main.csv          — one row per in-scope file (the deliverable)
  * delivery_excluded.csv      — enumerated-but-excluded non-fisheries TN/RL
  and, if --xlsx given, injects both as sheets into the checklist workbook.

Run:
    python tools/delivery_manifest.py --out ./wto_fish_out_v6 \
        --xlsx ./语料库数据确认清单.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from urllib.parse import quote

DM = Path("docs_manifest")
FE_S = "https://docs.wto.org/dol2fe/Pages/FE_Search/FE_S_S006.aspx"
FISH_RE = re.compile(
    r"fish|fisher|fishing|overfish|overcapacit|IUU|illegal[,\s].{0,20}unreported|"
    r"marine\s+capture|subsidies\s+(?:to|for)\s+fish", re.IGNORECASE)

# Site category -> short Chinese inclusion reason.
SITE_REASON = {
    "overview": "渔业补贴专题门户页（起始种子页）",
    "introduction": "专题页：渔业补贴谈判导论/背景",
    "legal_text": "渔业补贴协定正式法律文本",
    "ratification": "专题页：接受/批准协定（成员程序与清单）",
    "implementation": "专题页：协定履约报告",
    "fish_fund": "专题页：渔业基金（技援与能力建设）",
    "publication": "渔业补贴官方出版物（由专题页/出版库链接）",
    "news": "WTO 渔业补贴新闻稿（文件名含 fish）",
    "ministerial": "部长会渔业补贴简报",
    "case_story": "WTO 渔业补贴专题案例故事",
    "international_instrument": "门户页“相关国际文书”区转载的公约/准则",
    "multimedia": "专题页链接的渔业谈判视频",
}
CAT_ZH = {"overview": "概览", "introduction": "导论", "legal_text": "法律文本",
          "ratification": "接受与批准", "implementation": "履约", "fish_fund": "渔业基金",
          "publication": "出版物", "news": "新闻", "ministerial": "部长会简报",
          "case_story": "案例故事", "international_instrument": "国际文书",
          "multimedia": "音视频", "negotiation_submission": "谈判", "committee": "委员会",
          "mandate_decision": "部长决定与议定书", "uncategorized": "未分类"}


def load(p: Path) -> list[dict]:
    return [json.loads(l) for l in p.read_text(encoding="utf-8").splitlines()
            if l.strip()] if p.exists() else []


def catalogue_url(symbol: str | None) -> str:
    """A reproducible search URL that proves a symbol exists in the catalogue."""
    return (f"{FE_S}?Query={quote('(@Symbol= ' + (symbol or '') + ')')}"
            "&Language=ENGLISH&Context=FomerScriptedSearch&languageUIChanged=true")


def keywords(text: str) -> str:
    kws = sorted({m.lower() for m in FISH_RE.findall(text or "")})
    return "、".join(kws)


def _name(rec: dict) -> str:
    # listing 'text' begins with the symbol; strip it for a cleaner title
    t = re.sub(r"^\S+\s*", "", rec.get("text", "") or "").strip(" -;")
    return t[:140] or rec.get("symbol", "")


def build(out: Path) -> tuple[list[dict], list[dict]]:
    main: list[dict] = []
    excluded: list[dict] = []

    # ---- 1) site corpus (manifest) ----
    for r in load(out / "manifest.jsonl"):
        if r.get("duplicate_of"):
            continue
        err = r.get("error")
        blocked = err and "access-gated" in err
        if err and not blocked:
            continue
        cat = r.get("category", "uncategorized")
        if r.get("out_md_path"):
            status, local = "已转Markdown", r["out_md_path"]
        elif blocked:
            status, local = "受限·不可下载（需浏览器/会员）", ""
        elif r.get("content_type") == "pdf":
            status, local = "已下载·待解析", f"raw/pdf/{r.get('raw_sha256')}.pdf"
        else:
            status, local = "已下载", ""
        main.append({
            "类别/系列": CAT_ZH.get(cat, cat), "文件号": "",
            "名称/标题": (r.get("title") or r["url"].rsplit("/", 1)[-1])[:140],
            "状态": status, "原始URL": r["url"],
            "纳入/下载理由": SITE_REASON.get(cat, "渔业补贴专题站点文件"),
            "本地路径": local,
        })

    # ---- 2) manual additions (teacher-confirmed) ----
    for r in load(DM / "manual_additions.jsonl"):
        main.append({
            "类别/系列": CAT_ZH.get(r.get("category"), r.get("category", "")), "文件号": "",
            "名称/标题": r["name"], "状态": r.get("status", "已下载·待解析"),
            "原始URL": r.get("url", ""),
            "纳入/下载理由": r.get("note", "老师确认纳入"), "本地路径": r.get("raw_path", ""),
        })

    # ---- 3) document-library series ----
    SERIES = [
        ("TN/RL", DM / "tn_rl_listing.jsonl", True,  "规则谈判组文件"),
        ("RD/TN/RL", DM / "rd_tn_rl_listing.jsonl", False, "规则谈判室文件（非正式）"),
        ("JOB/RL", DM / "job_rl_listing.jsonl", False, "规则组室文件"),
        ("G/FS", DM / "gfs_listing.jsonl", True,  "渔业补贴委员会文件"),
    ]
    for label, listing, public, what in SERIES:
        for r in load(listing):
            url = r.get("english_url") or catalogue_url(r.get("symbol", ""))
            if not r.get("fisheries"):
                yr = re.findall(r"\b(19|20)\d{2}\b", r.get("text", ""))
                excluded.append({
                    "系列": label, "文件号": r.get("symbol", ""), "标题": _name(r),
                    "原始URL": url,
                    "排除理由": "规则谈判组非渔业文件（反倾销/SCM/区域协定等），标题无渔业字样"
                                + (f"；约 {yr[-1]}" if yr else ""),
                })
                continue
            if label == "G/FS":
                reason = "渔业补贴委员会专司渔业补贴，全系列纳入"
            else:
                reason = f"{what}，标题命中渔业关键词：{keywords(r.get('text',''))}"
            if r.get("downloaded"):
                status, local = "已下载·待解析", r.get("raw_path", "")
            elif public:
                status, local = "待下载（暂未提供PDF）", ""
            else:
                status, local = "受限·不可公开下载（仅目录可证存在）", ""
                reason += "；正文需 WTO 成员权限"
            main.append({
                "类别/系列": label, "文件号": r.get("symbol", ""),
                "名称/标题": _name(r), "状态": status, "原始URL": url,
                "纳入/下载理由": reason, "本地路径": local,
            })

    # ---- 4) ministerial / legal known decisions (from docs_manifest probe) ----
    # already covered by the series sweeps above (don't list twice)
    sweep_syms = {"TN/RL/31", "G/FS/1", "G/FS/W/1", "G/FS/M/1"}
    LEGAL_REASON = {
        "WT/MIN(22)/33": "MC12 部长决定：通过《渔业补贴协定》",
        "WT/MIN(17)/64": "MC11 部长决定：渔业补贴谈判授权",
        "WT/L/1144": "接受/修正议定书；不单独提供PDF，议定书内容已含于 WT/MIN(22)/33",
    }
    for r in load(DM / "docs_manifest.jsonl"):
        sym = r["symbol"]
        if sym in sweep_syms:
            continue
        if r.get("downloaded"):
            status, url, local = "已下载·待解析", r.get("url", ""), r.get("raw_path", "")
        else:
            status, url, local = "受限/未单独提供PDF（目录可证）", catalogue_url(sym), ""
        main.append({
            "类别/系列": "部长决定与议定书", "文件号": sym, "名称/标题": LEGAL_REASON.get(sym, sym),
            "状态": status, "原始URL": url,
            "纳入/下载理由": LEGAL_REASON.get(sym, "渔业补贴相关部长/法律文件"), "本地路径": local,
        })

    for i, row in enumerate(main, 1):
        row["序号"] = i
    return main, excluded


COLS_MAIN = ["序号", "类别/系列", "文件号", "名称/标题", "状态", "原始URL", "纳入/下载理由", "本地路径"]
COLS_EXCL = ["系列", "文件号", "标题", "原始URL", "排除理由"]


def write_csv(path: Path, rows: list[dict], cols: list[str]) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def inject_xlsx(xlsx: Path, main: list[dict], excluded: list[dict]) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    for name, rows, cols in [("交付总表（逐文件·带URL与理由）", main, COLS_MAIN),
                             ("排除清单（非渔业·已枚举）", excluded, COLS_EXCL)]:
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        ws.append(cols)
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
    wb.save(xlsx)


def main() -> int:
    ap = argparse.ArgumentParser(description="Per-file delivery manifest (URL + reason)")
    ap.add_argument("--out", default="./wto_fish_out_v6")
    ap.add_argument("--xlsx", help="also inject sheets into this checklist workbook")
    args = ap.parse_args()

    main_rows, excl_rows = build(Path(args.out))
    write_csv(DM / "delivery_main.csv", main_rows, COLS_MAIN)
    write_csv(DM / "delivery_excluded.csv", excl_rows, COLS_EXCL)
    if args.xlsx:
        inject_xlsx(Path(args.xlsx), main_rows, excl_rows)

    dl = sum(1 for r in main_rows if r["状态"].startswith("已"))
    restricted = sum(1 for r in main_rows if "受限" in r["状态"])
    print(f"交付主表: {len(main_rows)} 行（已下载/转换 {dl}，受限/待 {restricted}）")
    print(f"排除清单: {len(excl_rows)} 行")
    print(f"  -> {DM/'delivery_main.csv'}")
    print(f"  -> {DM/'delivery_excluded.csv'}")
    if args.xlsx:
        print(f"  -> 已注入 {args.xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
