"""Build the teacher-review deliverables from a crawl output dir.

Produces, from ``<out>/manifest.jsonl``:
  1. ``<out>/for_teacher/<中文类别>/<可读名>.<ext>`` — a browsable folder where
     every HTML item is its converted Markdown and every PDF is the raw file
     (pending parse). Access-gated videos get a note file (they cannot be
     downloaded via direct GET — see the crawler's blocked-media handling).
  2. An updated ``语料库数据确认清单.xlsx`` — the three WTO sheets are rewritten
     to match reality; the IOTC sheets are left untouched.

Run:
    python tools/build_review.py --out ./wto_fish_out_v6 \
        --xlsx ./语料库数据确认清单.xlsx
"""

from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path

import openpyxl

from wto_fish import classify

# Category code -> Chinese label (matches the existing checklist vocabulary).
CATEGORY_ZH: dict[str, str] = {
    "overview": "概览",
    "introduction": "导论",
    "legal_text": "法律文本",
    "ratification": "接受与批准",
    "implementation": "履约",
    "fish_fund": "渔业基金",
    "publication": "出版物",
    "news": "新闻",
    "ministerial": "部长会简报",
    "case_story": "案例故事",
    "international_instrument": "国际文书",
    "negotiation_submission": "谈判",
    "multimedia": "音视频",
    "committee": "委员会",
    "mandate_decision": "部长决定与议定书",
    "uncategorized": "未分类",
}
# Display order for the checklist / folders.
CATEGORY_ORDER = list(CATEGORY_ZH.keys())

# Human-readable names keyed by URL basename. Reused from the prior checklist;
# new items (recovered pages, videos) added here.
NAME_BY_BASENAME: dict[str, str] = {
    "fish_e.htm": "渔业补贴专题门户 / 协定概览",
    "fish_intro_e.htm": "渔业补贴谈判导论 / 背景",
    "agreement_fisheries_subsidies_e.htm": "如何接受议定书（成员程序指引）",
    "fish_acceptances_e.htm": "成员接受 / 批准清单",
    "implementfishagreement22_e.htm": "发展中 / 最不发达国家履约报告",
    "fish_fund_e.htm": "WTO 渔业基金（技援与能力建设）",
    "agreement_on_fisheries_subsidies_information_session_22_may_2025_final.pdf":
        "2025-05-22 信息通报会材料",
    "fishagree_e.htm": "协定宣传册（目标与预期收益）",
    "fish_factsheet_e.pdf": "渔业补贴情况说明书（Factsheet）",
    "fish_15sep25_e.htm": "协定生效新闻（2025-09-15）",
    "fish_10oct22_e.htm": "渔业补贴新闻（2022-10-10）",
    "fish_08nov22_e.htm": "渔业补贴新闻（2022-11-08）",
    "fish_14jun22_e.htm": "渔业补贴新闻（2022-06-14）",
    "fisheries_subsidies_e.htm": "MC13 部长会渔业补贴简报",
    "bffish_e.htm": "MC12 部长会渔业补贴简报",
    "stories_tanzania_e.htm": "坦桑尼亚案例",
    "stories-fijipacific_e.htm": "斐济 / 太平洋案例",
    "stories_from_the_ocean_e.htm": "海洋故事（总览页）",
    "stories-trinidadtobago_e.htm": "特立尼达和多巴哥案例",
    "stories_oman_e.htm": "阿曼案例",
    "stories-malaysia_e.htm": "马来西亚案例",
    "1969_vclt.pdf": "1969 维也纳条约法公约（VCLT）",
    "1982_unclos.pdf": "1982 联合国海洋法公约（UNCLOS）",
    "1995_fao_ccrf.pdf": "1995 FAO 负责任渔业行为守则（CCRF）",
    "1995_unfsa.pdf": "1995 联合国鱼类种群协定（UNFSA）",
    "2001_ipoa_iuu.pdf": "2001 FAO IUU 捕捞国际行动计划（IPOA-IUU）",
    "2009_psma.pdf": "2009 港口国措施协定（PSMA）",
    "2014_vg_fsp.pdf": "2014 FAO 船旗国表现自愿准则（VG-FSP）",
    "2015_fao_ssf.pdf": "2015 FAO 小规模渔业自愿准则（SSF）",
    "2017_vg_cds.pdf": "2017 FAO 渔获文件方案自愿准则（VG-CDS）",
    "ngr_presentation_on_rfmo_as.pdf": "规则谈判组 RFMO/As 演示材料",
    "2024_07_05_fish_chair_update.mp4": "主席渔业谈判进展视频（2024-07-05）",
    "fish_20240222.mp4": "MC13 渔业补贴视频（2024-02-22）",
    "fish_28022024.mp4": "MC13 渔业补贴视频（2024-02-28）",
}


# Overrides keyed by a URL substring, checked before the basename map. Needed
# where two pages share a basename (e.g. the overview and the legal text are
# both .../fish_e.htm).
NAME_BY_URL_SUBSTR: dict[str, str] = {
    "/docs_e/legal_e/fish_e.htm": "渔业补贴协定法律文本（正式条文）",
}


# Human-readable names for pending Tier-2 (docs.wto.org) documents.
DOC_NAME_BY_SYMBOL: dict[str, str] = {
    "WT/MIN(22)/33": "部长决定：通过《渔业补贴协定》（WT/MIN(22)/33）",
    "WT/MIN(17)/64": "MC11 渔业补贴部长决定（WT/MIN(17)/64）",
    "WT/L/1144": "接受议定书 / 修正议定书（WT/L/1144）",
    "TN/RL/31": "规则谈判组渔业补贴文件（TN/RL/31）",
    "G/FS/1": "渔业补贴委员会文件（G/FS/1）",
}
# Top-level folder under for_teacher/ for items pending the teacher's scope call.
PENDING_DIR = "待确认（请老师核定是否纳入）"


# The "Implementing the WTO Agreement on Fisheries Subsidies" publication
# (res_e/booksp_e), split into parts. Keyed by lowercased basename.
NAME_BY_BASENAME.update({
    "implementfishagreement22_e.pdf": "实施《渔业补贴协定》出版物（完整版）",
    "impfishag_dgmessage_e.pdf": "实施出版物·总干事致辞",
    "impfishag_part_1_e.pdf": "实施出版物·第一部分",
    "impfishag_part_2_e.pdf": "实施出版物·第二部分",
    "impfishag_part_3_e.pdf": "实施出版物·第三部分",
    "impfishag_conclusion_e.pdf": "实施出版物·结论",
    "impfishag_annex_how_to_e.pdf": "实施出版物·附录（如何接受协定）",
    "impfishag_faq_e.pdf": "实施出版物·常见问题",
    "impfishag_model_e.pdf": "实施出版物·示范接受文书",
    "fishagree_e.pdf": "协定宣传册（PDF 版）",
})
_NAME_BY_BASENAME_LC = {k.lower(): v for k, v in NAME_BY_BASENAME.items()}


def _basename(url: str) -> str:
    return url.rstrip("/").split("?", 1)[0].split("/")[-1]


def _name_for(url: str, base: str, title: str | None) -> str:
    for sub, name in NAME_BY_URL_SUBSTR.items():
        if sub in url:
            return name
    if base.lower() in _NAME_BY_BASENAME_LC:
        return _NAME_BY_BASENAME_LC[base.lower()]
    return title or base.rsplit(".", 1)[0]   # strip extension to avoid name.pdf.pdf


def _safe(name: str) -> str:
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip()


def _is_blocked(rec: dict) -> bool:
    return bool(rec.get("error") and "access-gated" in rec["error"])


def load_items(out: Path) -> list[dict]:
    """Return enriched, de-duplicated kept items from the manifest."""
    rows = [json.loads(l) for l in (out / "manifest.jsonl").read_text(
        encoding="utf-8").splitlines() if l.strip()]
    items = []
    for r in rows:
        if r.get("duplicate_of"):
            continue
        # Drop failed fetches (soft-404, HTTP errors). Keep access-gated media —
        # those are listed (as 受限) further down.
        if r.get("error") and not _is_blocked(r):
            continue
        url = r["url"]
        base = _basename(url)
        cat = classify.classify(url, r.get("title"))  # recompute with current rules
        if r.get("out_md_path"):
            kind, status, artifact = "HTML", "已转 Markdown", r["out_md_path"]
        elif _is_blocked(r):
            kind, status, artifact = "视频", "受限·需浏览器获取", None
        elif r.get("content_type") == "pdf":
            kind, status, artifact = "PDF", "已下载·待解析", f"raw/pdf/{r['raw_sha256']}.pdf"
        elif r.get("error"):
            kind, status, artifact = (r.get("content_type") or "?"), f"失败：{r['error'][:30]}", None
        else:
            kind, status, artifact = (r.get("content_type") or "?"), "已下载", None
        items.append({
            "url": url, "category": cat,
            "name": _name_for(url, base, r.get("title")),
            "kind": kind, "status": status, "artifact": artifact, "ext": base.rsplit(".", 1)[-1],
        })
    items.sort(key=lambda it: (CATEGORY_ORDER.index(it["category"])
                               if it["category"] in CATEGORY_ORDER else 999, it["name"]))
    return items


def load_manual(path: Path) -> list[dict]:
    """Manually-added, teacher-confirmed corpus items (e.g. foundational legal
    texts not reachable by the bounded crawl). Same shape as load_items()."""
    if not path.exists():
        return []
    out = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        rp = r["raw_path"]
        out.append({
            "url": r.get("url", ""), "category": r.get("category", "uncategorized"),
            "name": r["name"], "kind": r.get("kind", "PDF"),
            "status": r.get("status", "已下载·待解析"),
            "artifact": rp, "ext": rp.rsplit(".", 1)[-1],
        })
    return out


def load_tnrl(listing: Path) -> tuple[int, list[dict]]:
    """Return (total enumerated, fisheries-flagged records) from a TN/RL listing."""
    if not listing.exists():
        return 0, []
    rows = [json.loads(l) for l in listing.read_text(encoding="utf-8").splitlines() if l.strip()]
    fish = [r for r in rows if r.get("fisheries")]
    return len(rows), fish


def load_docs(docs_manifest: Path) -> list[dict]:
    """Pending Tier-2 documents that resolved via directdoc (downloaded)."""
    if not docs_manifest.exists():
        return []
    out = []
    for line in docs_manifest.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        r = json.loads(line)
        if not r.get("downloaded"):
            continue
        sym = r["symbol"]
        out.append({
            "symbol": sym, "series": r.get("series", ""),
            "category": classify.classify(sym),  # recompute with current rules
            "name": DOC_NAME_BY_SYMBOL.get(sym, sym),
            "raw_path": r["raw_path"], "size": r.get("size", 0),
        })
    return out


def build_folder(out: Path, items: list[dict], docs: list[dict],
                 tnrl_fish: list[dict] | None = None,
                 gfs_records: list[dict] | None = None) -> None:
    dest = out / "for_teacher"
    if dest.exists():
        # ignore_errors: an empty leftover dir held open by a file-explorer
        # window (Windows) can't be removed; that's harmless — files get
        # overwritten by name and the dir is recreated below.
        shutil.rmtree(dest, ignore_errors=True)
    tnrl_fish = tnrl_fish or []
    gfs_records = gfs_records or []
    # Symbols already covered by a full-series download: drop them from the
    # curated `docs` sample so they aren't filed twice.
    series_syms = {r.get("symbol") for r in (*tnrl_fish, *gfs_records) if r.get("raw_path")}
    docs = [d for d in docs if d["symbol"] not in series_syms]
    notes: dict[str, list[str]] = {}
    for it in items:
        zh = CATEGORY_ZH.get(it["category"], it["category"])
        folder = dest / zh
        folder.mkdir(parents=True, exist_ok=True)
        if it["artifact"]:
            src = out / it["artifact"]
            if src.exists():
                ext = "md" if it["artifact"].endswith(".md") else it["ext"]
                (folder / f"{_safe(it['name'])}.{ext}").write_bytes(src.read_bytes())
        else:
            notes.setdefault(zh, []).append(f"- {it['name']}  ——  {it['status']}\n  {it['url']}")
    for zh, lines in notes.items():
        (dest / zh / "_说明（未下载项）.txt").write_text(
            "以下条目无法直接下载（多为受访问限制的视频），需在浏览器中手动获取：\n\n"
            + "\n".join(lines) + "\n", encoding="utf-8")

    # Pending Tier-2 documents go under a clearly-separated folder so the
    # confirmed topic corpus above stays clean.
    for d in docs:
        zh = CATEGORY_ZH.get(d["category"], d["category"])
        folder = dest / PENDING_DIR / zh
        folder.mkdir(parents=True, exist_ok=True)
        (folder / f"{_safe(d['name'])}.pdf").write_bytes((out / d["raw_path"]).read_bytes())

    # Full series downloaded by docs_fetch --listing -> their category folders.
    def _place(records: list[dict], category: str) -> int:
        folder = dest / PENDING_DIR / CATEGORY_ZH[category]
        n = 0
        for r in records:
            src = out / r["raw_path"] if r.get("raw_path") else None
            if not src or not src.exists():
                continue
            folder.mkdir(parents=True, exist_ok=True)
            (folder / f"{_safe(r.get('symbol') or 'doc')}.pdf").write_bytes(src.read_bytes())
            n += 1
        return n

    n_tnrl = _place(tnrl_fish, "negotiation_submission")
    n_gfs = _place(gfs_records, "committee")

    if docs or n_tnrl or n_gfs:
        (dest / PENDING_DIR / "_读我.txt").write_text(
            "本目录是 docs.wto.org 文档库（部长决定 / 谈判提案 / 委员会文件）中\n"
            "经 directdoc 匿名下载的英文文档，已分类放好，但【是否纳入本专题待老师确认】。\n"
            "均未解析为 Markdown。文件名为文档号，标题对照见清单对应枚举表。\n\n"
            f"· 谈判/：TN/RL 系列中标题含渔业关键词的 {n_tnrl} 份（见“TN-RL枚举(渔业)”表）。\n"
            f"· 委员会/：渔业补贴委员会 G/FS 全系列 {n_gfs} 份（见“G-FS枚举”表）。\n"
            "· 部长决定与议定书/：核心部长决定样本。\n",
            encoding="utf-8")


def _clear(ws) -> None:
    ws.delete_rows(1, ws.max_row or 1)


def update_xlsx(xlsx: Path, items: list[dict], docs: list[dict],
                tnrl_total: int = 0, tnrl_fish: list[dict] | None = None,
                gfs_total: int = 0, gfs_records: list[dict] | None = None) -> None:
    wb = openpyxl.load_workbook(xlsx)
    tnrl_fish = tnrl_fish or []
    gfs_records = gfs_records or []
    n_tnrl_dl = sum(1 for r in tnrl_fish if r.get("downloaded"))
    n_gfs_dl = sum(1 for r in gfs_records if r.get("downloaded"))
    n_doc = {s: sum(1 for d in docs if d["series"] == s)
             for s in ("WT/MIN", "WT/L", "TN/RL", "G/FS")}
    instruments = "、".join(it["name"] for it in items
                           if it["category"] == "international_instrument") or "（无）"

    ws = wb["WTO已爬取清单"]
    _clear(ws)
    ws.append(["序号", "类别", "资料名称", "类型", "状态", "来源 URL"])
    for i, it in enumerate(items, 1):
        ws.append([i, CATEGORY_ZH.get(it["category"], it["category"]),
                   it["name"], it["kind"], it["status"], it["url"]])

    ws2 = wb["WTO分类汇总"]
    _clear(ws2)
    ws2.append(["类别", "数量"])
    counts: dict[str, int] = {}
    for it in items:
        zh = CATEGORY_ZH.get(it["category"], it["category"])
        counts[zh] = counts.get(zh, 0) + 1
    for zh in [CATEGORY_ZH[c] for c in CATEGORY_ORDER if CATEGORY_ZH[c] in counts]:
        ws2.append([zh, counts[zh]])
    ws2.append(["合计", len(items)])

    ws3 = wb["WTO待确认范围"]
    _clear(ws3)
    ws3.append(["类别", "拟纳入内容", "获取方式 / 成本 / 验证结果", "是否属于本专题（请老师确认）"])
    ws3.append(["文档库·部长决定/议定书",
                "WT/MIN(22)/33（通过协定的部长决定）、WT/MIN(17)/64（MC11 决定）、"
                "WT/L/1144（接受/修正议定书）等",
                f"✓ directdoc 匿名直取可用。已下载 {n_doc['WT/MIN'] + n_doc['WT/L']} 份"
                "（WT/MIN(22)/33、WT/MIN(17)/64）放入 for_teacher/待确认/；"
                "WT/L/1144 直取未命中，需手动核实文档号", ""])
    ws3.append(["文档库·谈判提案",
                "规则谈判组谈判提案（TN/RL 系列，2001 年至今）",
                f"✓ 脚本检索枚举全系列共 {tnrl_total or '—'} 份，其中标题含渔业关键词 {len(tnrl_fish)} 份。"
                f"这 {len(tnrl_fish)} 份英文版【已全部下载 {n_tnrl_dl} 份】放入 "
                "for_teacher/待确认/谈判/（文件名=文档号，标题对照见“TN-RL枚举(渔业)”表），未解析为 Markdown，待老师确认纳入。"
                "注：标题无“渔业”字样的会议纪要等程序性文件未计入", ""])
    ws3.append(["文档库·委员会文件",
                "渔业补贴委员会文件（协定 2025-09-15 生效后新设，G/FS/ 系列已确认）",
                f"✓ 系列号确认为 G/FS/。已枚举全系列共 {gfs_total or len(gfs_records)} 份，"
                f"英文版【已下载 {n_gfs_dl} 份】放入 for_teacher/待确认/委员会/（见“G-FS枚举”表 / "
                "docs_manifest/gfs_listing.jsonl）。多为成员履约/通报文件，未解析为 Markdown，待老师确认纳入。"
                "该委员会将持续产出新文件，需定期重跑枚举", ""])
    ws3.append(["音视频·受限",
                "渔业谈判主席视频、MC13 视频等（.mp4，共 3 个）",
                "⚠ 直接 GET 被重定向到登录/错误页，无法匿名下载；需浏览器/流式抓取。已在清单中标记", ""])
    ws3.append(["国际文书·归属待定",
                instruments,
                "已下载（9 份，见清单“国际文书”类）；均为 WTO 转载的外部公约/准则，"
                "是否计入本专题请老师确认", ""])

    # Detail of the pending Tier-2 downloads (separate sheet so it doesn't
    # inflate the confirmed-corpus checklist).
    name = "WTO待确认下载明细"
    if name in wb.sheetnames:
        del wb[name]
    ws4 = wb.create_sheet(name)
    ws4.append(["文档号", "系列", "类别", "大小(KB)", "状态", "本地路径"])
    for d in docs:
        ws4.append([d["symbol"], d["series"], CATEGORY_ZH.get(d["category"], d["category"]),
                    round(d["size"] / 1024, 1), "已下载·待老师确认纳入",
                    f"for_teacher/{PENDING_DIR}/{CATEGORY_ZH.get(d['category'], d['category'])}/"
                    f"{_safe(d['name'])}.pdf"])

    # TN/RL fisheries enumeration (listing only; teacher picks before download).
    tname = "TN-RL枚举(渔业)"
    if tname in wb.sheetnames:
        del wb[tname]
    if tnrl_fish:
        ws5 = wb.create_sheet(tname)
        ws5.append(["文档号", "标题 / 会议信息", "英文 directdoc 链接", "状态", "本地路径"])
        tn_zh = CATEGORY_ZH["negotiation_submission"]
        for r in tnrl_fish:
            dl = r.get("downloaded")
            local = (f"for_teacher/{PENDING_DIR}/{tn_zh}/{_safe(r.get('symbol') or '')}.pdf"
                     if dl else "")
            ws5.append([r.get("symbol", ""), r.get("text", ""), r.get("english_url", ""),
                        "已下载·待确认纳入" if dl else "未下载", local])

    # G/FS committee enumeration.
    gname = "G-FS枚举"
    if gname in wb.sheetnames:
        del wb[gname]
    if gfs_records:
        ws6 = wb.create_sheet(gname)
        ws6.append(["文档号", "标题 / 会议信息", "英文 directdoc 链接", "状态", "本地路径"])
        gfs_zh = CATEGORY_ZH["committee"]
        for r in gfs_records:
            dl = r.get("downloaded")
            local = f"for_teacher/{PENDING_DIR}/{gfs_zh}/{_safe(r.get('symbol') or '')}.pdf" if dl else ""
            ws6.append([r.get("symbol", ""), r.get("text", ""), r.get("english_url", ""),
                        "已下载·待确认纳入" if dl else "未下载", local])

    wb.save(xlsx)


def main() -> int:
    ap = argparse.ArgumentParser(description="Build teacher-review folder + update checklist xlsx")
    ap.add_argument("--out", required=True, help="crawl output dir (has manifest.jsonl)")
    ap.add_argument("--xlsx", required=True, help="checklist .xlsx to update in place")
    ap.add_argument("--docs-manifest", default="./docs_manifest/docs_manifest.jsonl",
                    help="Tier-2 pending-docs manifest (from docs_fetch.py)")
    ap.add_argument("--tnrl-listing", default="./docs_manifest/tn_rl_listing.jsonl",
                    help="TN/RL enumeration (from docs_enumerate.py)")
    ap.add_argument("--gfs-listing", default="./docs_manifest/gfs_listing.jsonl",
                    help="G/FS committee enumeration (from docs_enumerate.py)")
    ap.add_argument("--manual", default="./docs_manifest/manual_additions.jsonl",
                    help="manually-added, teacher-confirmed corpus items")
    args = ap.parse_args()

    out = Path(args.out)
    items = load_items(out) + load_manual(Path(args.manual))
    items.sort(key=lambda it: (CATEGORY_ORDER.index(it["category"])
                               if it["category"] in CATEGORY_ORDER else 999, it["name"]))
    docs = load_docs(Path(args.docs_manifest))
    tnrl_total, tnrl_fish = load_tnrl(Path(args.tnrl_listing))
    gfs_total, gfs_records = load_tnrl(Path(args.gfs_listing))
    build_folder(out, items, docs, tnrl_fish, gfs_records)
    update_xlsx(Path(args.xlsx), items, docs, tnrl_total, tnrl_fish, gfs_total, gfs_records)

    # Standalone CSVs (robust fallback, openable without Excel lock issues).
    import csv

    def _csv(records, path):
        with Path(path).open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["symbol", "title", "english_directdoc_url", "downloaded"])
            for r in records:
                w.writerow([r.get("symbol", ""), r.get("text", ""),
                            r.get("english_url", ""), bool(r.get("downloaded"))])
    if tnrl_fish:
        _csv(tnrl_fish, Path(args.tnrl_listing).with_name("tn_rl_fisheries.csv"))
    if gfs_records:
        _csv(gfs_records, Path(args.gfs_listing).with_name("gfs_listing.csv"))

    print(f"items: {len(items)} | pending docs: {len(docs)} | "
          f"TN/RL enumerated: {tnrl_total} (fisheries {len(tnrl_fish)}) | "
          f"G/FS: {gfs_total or len(gfs_records)}")
    print(f"for_teacher/: {out / 'for_teacher'}")
    print(f"xlsx updated: {args.xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
