"""Crawl the country-profile pages of members that accepted the Fisheries
Subsidies Agreement (scheme C).

Country pages (/english/thewto_e/countries_e/) are outside the bounded fisheries
crawl, but the teacher asked to include the accepting members' profiles. The
links are JS-rendered, so we resolve them from the rendered WTO members
directory (docs_manifest/members_map.json) matched against the 94 acceptance
rows. The country pages themselves are static, so a plain fetch + Markdown is
enough.

Outputs:
  接受成员国家页/<member>.md            one Markdown per country profile
  接受成员国家页/_接受成员国家页索引.{xlsx,csv}  member | date | title | url | path

Run:
    python tools/crawl_countries.py
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import time
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill

from wto_fish import extract

ROOT = Path("接受成员国家页")
UA = "wto-fish-corpus-bot/1.0 (research; contact: zhihao7946@gmail.com)"


def _safe(name: str) -> str:
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip()[:80]


def acceptance_dates(site_out: Path) -> dict[str, str]:
    """member -> acceptance date, parsed from the rendered acceptances table."""
    dates = {}
    for f in (site_out / "markdown").glob("*.md"):
        t = f.read_text(encoding="utf-8")
        if "Acceptance date" not in t:
            continue
        for m in re.finditer(r"^\|\s*([^|]+?)\s*\|\s*(\d+\s+\w+\s+\d{4})\s*\|", t, re.M):
            dates[m.group(1).strip()] = m.group(2).strip()
        break
    return dates


def main() -> int:
    ap = argparse.ArgumentParser(description="Crawl accepting members' country pages")
    ap.add_argument("--map", default="./docs_manifest/accepting_country_pages.json")
    ap.add_argument("--site-out", default="./site_out")
    ap.add_argument("--delay", type=float, default=0.6)
    args = ap.parse_args()

    members = json.loads(Path(args.map).read_text(encoding="utf-8"))   # member -> url
    dates = acceptance_dates(Path(args.site_out))
    ROOT.mkdir(parents=True, exist_ok=True)

    rows = []
    with httpx.Client(headers={"User-Agent": UA}, follow_redirects=True,
                      timeout=40.0, trust_env=False) as c:
        for i, (member, url) in enumerate(sorted(members.items()), 1):
            clean = re.sub(r"\s*\(See text.*$", "", member).strip()
            date = dates.get(member) or dates.get(clean, "")
            local = title = ""
            status = "已转 Markdown"
            try:
                r = c.get(url)
                if r.status_code == 200 and "html" in r.headers.get("content-type", "").lower():
                    md = extract.extract_markdown(r.text, url) or ""
                    h1 = re.search(r"^#\s+(.+)$", md, re.M)
                    title = h1.group(1).strip() if h1 else clean
                    path = ROOT / f"{_safe(clean)}.md"
                    front = (f"---\nmember: \"{clean}\"\nacceptance_date: \"{date}\"\n"
                             f"url: \"{url}\"\nsource: \"wto-country-profile\"\n---\n\n")
                    path.write_text(front + md, encoding="utf-8")
                    local = str(path).replace("\\", "/")
                else:
                    status = f"失败 HTTP {r.status_code}"
            except httpx.HTTPError as e:
                status = f"失败 {type(e).__name__}"
            rows.append({"成员": clean, "接受日期": date, "标题": title,
                         "原始URL": url, "本地路径": local, "状态": status})
            if i % 20 == 0 or i == len(members):
                print(f"  {i}/{len(members)}  ok={sum(1 for x in rows if x['本地路径'])}")
            time.sleep(args.delay)

    cols = ["序号", "成员", "接受日期", "标题", "原始URL", "本地路径", "状态"]
    with (ROOT / "_接受成员国家页索引.csv").open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f); w.writerow(cols)
        for i, r in enumerate(rows, 1):
            w.writerow([i] + [r[c] for c in cols[1:]])
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "接受成员国家页"
    ws.append(cols)
    for ce in ws[1]:
        ce.font = Font(bold=True, color="FFFFFF"); ce.fill = PatternFill("solid", fgColor="4472C4")
    for i, r in enumerate(rows, 1):
        ws.append([i, r["成员"], r["接受日期"], r["标题"], r["原始URL"], r["本地路径"], r["状态"]])
        lk = ws.cell(ws.max_row, 5); lk.hyperlink = r["原始URL"]
        lk.font = Font(color="0563C1", underline="single")
        if not r["本地路径"]:
            for ce in ws[ws.max_row]:
                ce.fill = PatternFill("solid", fgColor="FCE4E4")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:G{ws.max_row}"
    for col, w_ in zip("ABCDEFG", [5, 28, 16, 34, 58, 46, 14]):
        ws.column_dimensions[col].width = w_
    wb.save(ROOT / "_接受成员国家页索引.xlsx")

    ok = sum(1 for r in rows if r["本地路径"])
    print(f"\n国家页: {ok}/{len(rows)} 抓取成功 -> {ROOT}/ + _接受成员国家页索引.xlsx/.csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
