"""Finalize the checklist workbook: prune redundant sheets, add an index, reorder.

Run this LAST (after build_review.py and delivery_manifest.py). It enforces a
clean, teacher-facing structure and is idempotent:

  * drops sheets now fully contained in 交付总表 (待确认下载明细 / G-FS枚举 / TN-RL枚举)
  * (re)builds a 索引 cover sheet describing every sheet + scope + key numbers
  * orders the WTO sheets logically and leaves any other sheets (IOTC) untouched
    at the end.

Run:
    python tools/finalize_workbook.py --xlsx ./语料库数据确认清单.xlsx
"""

from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path

import openpyxl

INDEX = "索引"
DELIVERY = "交付总表（逐文件·带URL与理由）"
EXCLUDED = "排除清单（非渔业·已枚举）"

# Redundant now that 交付总表 holds every file with URL + reason.
REMOVE = ["WTO待确认下载明细", "G-FS枚举", "TN-RL枚举(渔业)"]
# Desired leading order for WTO sheets; unknown sheets (IOTC) keep trailing.
WTO_ORDER = [INDEX, "WTO已爬取清单", "WTO分类汇总", "WTO待确认范围", DELIVERY, EXCLUDED]


def _delivery_stats(wb) -> tuple[int, int, int, dict]:
    if DELIVERY not in wb.sheetnames:
        return 0, 0, 0, {}
    ws = wb[DELIVERY]
    hdr = [c.value for c in ws[1]]
    si = hdr.index("状态")
    statuses = [(r[si].value or "") for r in ws.iter_rows(min_row=2)]
    total = len(statuses)
    dl = sum(1 for s in statuses if s.startswith("已"))
    restricted = total - dl
    by = Counter(s.split("·")[0].split("（")[0] for s in statuses)
    return total, dl, restricted, by


def _excluded_count(wb) -> int:
    return wb[EXCLUDED].max_row - 1 if EXCLUDED in wb.sheetnames else 0


def build_index(wb) -> None:
    total, dl, restricted, by = _delivery_stats(wb)
    excl = _excluded_count(wb)
    if INDEX in wb.sheetnames:
        del wb[INDEX]
    ws = wb.create_sheet(INDEX, 0)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 92

    def row(a="", b=""):
        ws.append([a, b])

    row("WTO 渔业补贴语料 — 数据确认清单（索引）", "")
    row("", "")
    row("范围定义", "①渔业补贴专题站点（www.wto.org 渔业目录，有界爬取）"
        " ＋ ②docs.wto.org 上与渔业补贴相关的文档系列（TN/RL、RD/TN/RL、JOB/RL、G/FS）"
        " ＋ ③部长决定/议定书 ＋ ④老师确认的相关法律文本。")
    row("关键数字（在范围内）", f"共 {total} 份：已下载/已转换 {dl} 份；受限或暂无PDF {restricted} 份"
        f"（每份均附官方URL证明存在）。另排除非渔业 {excl} 份（见排除清单）。")
    if by:
        row("按状态", "　".join(f"{k}:{v}" for k, v in by.most_common()))
    row("", "")
    row("【各表说明】", "")
    row("WTO已爬取清单", "站点已爬取的核心文件（中文命名，便于阅读）。")
    row("WTO分类汇总", "站点文件按类别计数。")
    row("WTO待确认范围", "需老师/领导确认是否纳入的范围说明与验证结果。")
    row(DELIVERY, "★主交付：每份在范围内的文件一行，含【原始URL】（可点开核对存在）与【纳入/下载理由】"
        "（渔业文件标明命中的关键词）。")
    row(EXCLUDED, "已枚举但属非渔业、不纳入的文件，附URL与排除理由——证明“看过全集、排除有据”。")
    row("IOTC*", "另一项目数据，保留备查。")
    row("", "")
    row("【如何独立核验】", "")
    row("1. 计数可复现", "去 docs.wto.org 高级检索按符号（如 TN/RL/*）搜索，官网总数与我方枚举一致。")
    row("2. 逐条可点验", "交付总表每行的原始URL可直接打开核对；详见 docs_manifest/coverage_report.md。")
    row("3. 代码可重跑", "tools/ 下脚本一条命令重现同一结果。")


def main() -> int:
    ap = argparse.ArgumentParser(description="Finalize checklist workbook structure")
    ap.add_argument("--xlsx", required=True)
    args = ap.parse_args()
    xlsx = Path(args.xlsx)

    wb = openpyxl.load_workbook(xlsx)
    for name in REMOVE:
        if name in wb.sheetnames:
            del wb[name]
    build_index(wb)

    # Reorder: known WTO sheets first (in WTO_ORDER), then the rest as-is.
    lead = [wb[n] for n in WTO_ORDER if n in wb.sheetnames]
    rest = [ws for ws in wb._sheets if ws not in lead]
    wb._sheets = lead + rest

    wb.save(xlsx)
    print("最终表顺序：")
    for i, n in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
